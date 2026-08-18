import importlib.util
import io
import os
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SPEC = importlib.util.spec_from_file_location("app_under_test", ROOT / "app.py")
APP = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(APP)

os.environ["FRESH_FROZEN_SKIP_UI"] = "1"
FRESH_SPEC = importlib.util.spec_from_file_location(
    "fresh_frozen_app_under_test",
    next(path for path in ROOT.glob("app_*.py") if path.name != "app.py"),
)
FRESH_APP = importlib.util.module_from_spec(FRESH_SPEC)
FRESH_SPEC.loader.exec_module(FRESH_APP)


class AppHelperRegressionTests(unittest.TestCase):
    def test_byproduct_factor_map_loads_material_coefficients_from_workbook(self):
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            pd.DataFrame(
                {
                    "\u7269\u6599\u53f7": ["39001129", "39000416", "39000027"],
                    "\u7269\u6599\u63cf\u8ff0": ["a", "b", "c"],
                    "\u7cfb\u6570": [0.50, "70%", 0.65],
                }
            ).to_excel(writer, index=False)
        buf.seek(0)

        factors = APP._load_byproduct_factor_map(buf)

        self.assertEqual(factors["39001129"], 0.50)
        self.assertEqual(factors["39000416"], 0.70)
        self.assertEqual(factors["39000027"], 0.65)

    def test_market_price_map_loads_by_header_names(self):
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            pd.DataFrame(
                {
                    "\u5f53\u524d\u884c\u60c5\u4ef7": [12.5, 8.2],
                    "\u5206\u7c7b": ["\u817f\u8089", "\u80f8\u8089"],
                    "\u5de5\u5382": ["BB2", "BB2"],
                    "\u57fa\u671f\u884c\u60c5\u4ef7": [10.5, 7.1],
                }
            ).to_excel(writer, index=False, sheet_name="\u884c\u60c5\u4ef7\u8986\u76d6")
        buf.seek(0)

        prices = APP._load_market_price_map(buf)

        self.assertEqual(prices[("BB2", "\u817f\u8089")]["\u57fa\u671f\u884c\u60c5\u4ef7"], 10.5)
        self.assertEqual(prices[("BB2", "\u817f\u8089")]["\u5f53\u524d\u884c\u60c5\u4ef7"], 12.5)
        self.assertEqual(prices[("BB2", "\u80f8\u8089")]["\u57fa\u671f\u884c\u60c5\u4ef7"], 7.1)

    def test_market_plant_code_from_filename_maps_chinese_factory_names(self):
        cases = {
            "\u868c\u57e0\u4e00\u5382\u7cfb\u7edf\u6210\u672c-2604.xlsx": "BB1",
            "\u868c\u57e0\u4e8c\u5382\u7cfb\u7edf\u6210\u672c-2604.xlsx": "BB2",
            "\u5927\u8fde\u7cfb\u7edf\u6210\u672c-2604.xlsx": "DL",
            "\u8fbd\u9633\u7cfb\u7edf\u6210\u672c-2604.xlsx": "LY",
            "\u5929\u6d25\u7cfb\u7edf\u6210\u672c-2604.xlsx": "TJ",
            "\u5156\u5dde\u7cfb\u7edf\u6210\u672c-2604.xlsx": "YZ",
        }

        for filename, expected in cases.items():
            with self.subTest(filename=filename):
                self.assertEqual(APP._market_plant_code_from_filename(filename), expected)

        self.assertIsNone(APP._market_plant_code_from_filename("unknown.xlsx"))

    def test_market_plant_code_from_filenames_falls_back_to_other_upload_names(self):
        self.assertEqual(
            APP._market_plant_code_from_filenames(
                "BB_2604.xlsx",
                "\u539f\u6599\u6e05\u5355_\u868c\u57e0\u4e00\u5382.xlsx",
                "\u868c\u57e0\u4e00\u5382\u7cfb\u7edf\u6210\u672c-Q4(1).xlsx",
            ),
            "BB1",
        )
        self.assertEqual(
            APP._market_plant_code_from_filenames(
                "BB_2604.xlsx",
                "\u539f\u6599\u6e05\u5355_\u868c\u57e0\u4e8c\u5382.xlsx",
            ),
            "BB2",
        )

    def test_market_price_sheets_export_as_values_when_uploaded(self):
        mat_col = "\u7269\u6599\u53f7"
        desc_col = "\u7269\u6599\u63cf\u8ff0"
        row_col = "\u884c\u7c7b\u578b"
        impact_col = "\u5f71\u54cd\u53e3\u5f84"
        category_col = "\u5206\u7c7b"
        unit_col = "\u4fee\u5f62\u524d\u539f\u6599\u7efc\u5408\u8017\u7528\u5355\u4ef7"
        util_col = "\u4fee\u5f62\u5229\u7528\u7387"
        loss_col = "\u635f\u8017\u7387"
        raw_cost_col = "\u534a\u6210\u54c1\u539f\u6599\u6210\u672c"
        labor_col = "\u534a\u6210\u54c1\u4fee\u5f62\u4eba\u5de5\u6210\u672c"
        total_col = "\u534a\u6210\u54c1\u603b\u6210\u672c"
        in_qty_col = "\u534a\u6210\u54c1\u5165\u5e93\u91cf"
        bom_col = "BOM"
        bom_ratio_col = "BOM\u5360\u6bd4"
        raw_code = "31000001"
        comp_col = "\u7efc\u5408\u5355\u4ef7"

        def tsc_rows(mat, desc, category):
            labels = [
                "4\u6708\u5b9e\u9645\u5355\u4ef7",
                "25\u5e74\u5b9e\u9645\u5355\u4ef7",
                "4\u6708\u89c4\u683c\u5360\u6bd4",
                "\u89c4\u683c\u5360\u6bd4",
                "\u5dee\u5f02",
                "\u5bf9\u534a\u6210\u54c1\u6210\u672c\u7684\u5f71\u54cd",
                "\u5bf9\u534a\u6210\u54c1\u6210\u672c\u7684\u5f71\u54cd",
            ]
            impacts = ["", "", "", "", "", "\u5355\u4f4d\u6210\u672c", "\u603b\u6210\u672c"]
            return pd.DataFrame(
                [
                    {
                        mat_col: mat,
                        desc_col: desc,
                        row_col: label,
                        impact_col: impacts[i],
                        category_col: category,
                        unit_col: 9.0,
                        util_col: 0.8,
                        loss_col: 0.1,
                        raw_cost_col: 9.0,
                        labor_col: 1.0,
                        total_col: 10.0,
                        in_qty_col: 2.0,
                        bom_col: "",
                        bom_ratio_col: None,
                    }
                    for i, label in enumerate(labels)
                ]
            )

        leg = tsc_rows("39000001", "\u817f\u8089\u89c4\u683c", "\u817f\u8089")
        breast = tsc_rows("39000002", "\u80f8\u8089\u89c4\u683c", "\u80f8\u8089")
        empty = leg.iloc[0:0].copy()
        raw_usage_leg = leg[[mat_col, desc_col, row_col, category_col]].iloc[:5].copy()
        raw_usage_leg[raw_code] = [9.0, 8.0, 0.5, 0.4, 1.0]
        raw_usage_leg[comp_col] = [9.0, 8.0, 0.5, 0.4, 1.0]
        raw_usage_breast = breast[[mat_col, desc_col, row_col, category_col]].iloc[:5].copy()
        raw_usage_breast[raw_code] = [7.0, 6.0, 0.5, 0.4, 1.0]
        raw_usage_breast[comp_col] = [7.0, 6.0, 0.5, 0.4, 1.0]
        market = io.BytesIO()
        with pd.ExcelWriter(market, engine="openpyxl") as writer:
            pd.DataFrame(
                {
                    "\u5de5\u5382": ["BB2", "BB2"],
                    "\u5206\u7c7b": ["\u817f\u8089", "\u80f8\u8089"],
                    "\u57fa\u671f\u884c\u60c5\u4ef7": [10.0, 7.0],
                    "\u5f53\u524d\u884c\u60c5\u4ef7": [12.0, 8.0],
                }
            ).to_excel(writer, index=False, sheet_name="\u884c\u60c5\u4ef7\u8986\u76d6")
        market.seek(0)

        data = APP.to_excel_bytes(
            leg,
            breast,
            empty,
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame(),
            raw_usage_leg,
            raw_usage_breast,
            raw_usage_leg.iloc[0:0].copy(),
            {raw_code: "\u539f\u6599\u89c4\u683c"},
            {raw_code: "\u539f\u6599\u89c4\u683c"},
            {},
            "BB2",
            pd.DataFrame(),
            pd.DataFrame(),
            "4\u6708",
            {},
            market,
            quarter_label="Q4",
        )

        wb = load_workbook(io.BytesIO(data), data_only=False)
        self.assertIn("\u817f\u8089\u884c\u60c5-\u8f83\u5b63\u5ea6", wb.sheetnames)
        self.assertIn("\u80f8\u8089\u884c\u60c5-\u8f83\u5b63\u5ea6", wb.sheetnames)
        ws = wb["\u817f\u8089\u884c\u60c5-\u8f83\u5b63\u5ea6"]
        headers = [cell.value for cell in ws[2]]
        comp_idx = headers.index(comp_col) + 1
        row_idx = headers.index(row_col) + 1
        unit_idx = headers.index(unit_col) + 1
        util_idx = headers.index(util_col) + 1
        loss_idx = headers.index(loss_col) + 1
        raw_cost_idx = headers.index(raw_cost_col) + 1
        labor_idx = headers.index(labor_col) + 1
        total_idx = headers.index(total_col) + 1
        raw_idx = next(i + 1 for i, value in enumerate(headers) if APP._normalize_mat(value) == raw_code)
        factor = 0.95
        expected_current_raw = (12.0 - (1 - 0.8 - 0.1) * 12.0 * factor) / 0.8
        expected_ref_raw = (10.0 - (1 - 0.8 - 0.1) * 10.0 * factor) / 0.8
        expected_raw_diff = expected_current_raw - expected_ref_raw
        self.assertEqual(ws.cell(row=5, column=comp_idx).value, 12.0)
        self.assertEqual(ws.cell(row=6, column=row_idx).value, "25\u5e74Q4\u5b9e\u9645\u5355\u4ef7")
        self.assertEqual(ws.cell(row=6, column=comp_idx).value, 8.0)
        self.assertEqual(ws.cell(row=6, column=unit_idx).value, 10.0)
        self.assertEqual(ws.cell(row=6, column=util_idx).value, 0.8)
        self.assertEqual(ws.cell(row=6, column=loss_idx).value, 0.1)
        self.assertEqual(ws.cell(row=6, column=labor_idx).value, 1.0)
        self.assertAlmostEqual(ws.cell(row=6, column=raw_cost_idx).value, expected_ref_raw)
        self.assertAlmostEqual(ws.cell(row=6, column=total_idx).value, expected_ref_raw + 1.0)
        self.assertAlmostEqual(ws.cell(row=9, column=unit_idx).value, 2.0)
        self.assertAlmostEqual(ws.cell(row=9, column=raw_cost_idx).value, expected_raw_diff)
        self.assertAlmostEqual(ws.cell(row=10, column=unit_idx).value, expected_raw_diff)
        self.assertAlmostEqual(ws.cell(row=11, column=total_idx).value, expected_raw_diff * 2.0)
        self.assertTrue(ws.column_dimensions[ws.cell(row=2, column=raw_idx).column_letter].hidden)
        for sheet_name in ["\u817f\u8089\u884c\u60c5-\u8f83\u5b63\u5ea6", "\u80f8\u8089\u884c\u60c5-\u8f83\u5b63\u5ea6"]:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    self.assertFalse(isinstance(cell.value, str) and cell.value.startswith("="))

    def test_market_price_export_uses_market_plant_code_not_display_prefix(self):
        mat_col = "\u7269\u6599\u53f7"
        desc_col = "\u7269\u6599\u63cf\u8ff0"
        row_col = "\u884c\u7c7b\u578b"
        impact_col = "\u5f71\u54cd\u53e3\u5f84"
        category_col = "\u5206\u7c7b"
        unit_col = "\u4fee\u5f62\u524d\u539f\u6599\u7efc\u5408\u8017\u7528\u5355\u4ef7"
        util_col = "\u4fee\u5f62\u5229\u7528\u7387"
        loss_col = "\u635f\u8017\u7387"
        raw_cost_col = "\u534a\u6210\u54c1\u539f\u6599\u6210\u672c"
        labor_col = "\u534a\u6210\u54c1\u4fee\u5f62\u4eba\u5de5\u6210\u672c"
        total_col = "\u534a\u6210\u54c1\u603b\u6210\u672c"
        in_qty_col = "\u534a\u6210\u54c1\u5165\u5e93\u91cf"
        comp_col = "\u7efc\u5408\u5355\u4ef7"
        raw_code = "31000001"
        labels = [
            "4\u6708\u5b9e\u9645\u5355\u4ef7",
            "25\u5e74\u5b9e\u9645\u5355\u4ef7",
            "4\u6708\u89c4\u683c\u5360\u6bd4",
            "\u89c4\u683c\u5360\u6bd4",
            "\u5dee\u5f02",
            "\u5bf9\u534a\u6210\u54c1\u6210\u672c\u7684\u5f71\u54cd",
            "\u5bf9\u534a\u6210\u54c1\u6210\u672c\u7684\u5f71\u54cd",
        ]
        impacts = ["", "", "", "", "", "\u5355\u4f4d\u6210\u672c", "\u603b\u6210\u672c"]
        leg = pd.DataFrame(
            [
                {
                    mat_col: "39000308",
                    desc_col: "\u817f\u8089\u89c4\u683c",
                    row_col: label,
                    impact_col: impacts[i],
                    category_col: "\u817f\u8089",
                    unit_col: 9.0,
                    util_col: 0.8,
                    loss_col: 0.1,
                    raw_cost_col: 9.0,
                    labor_col: 1.0,
                    total_col: 10.0,
                    in_qty_col: 2.0,
                }
                for i, label in enumerate(labels)
            ]
        )
        empty = leg.iloc[0:0].copy()
        raw_usage_leg = leg[[mat_col, desc_col, row_col, category_col]].iloc[:5].copy()
        raw_usage_leg[raw_code] = [9.0, 8.0, 0.5, 0.4, 1.0]
        raw_usage_leg[comp_col] = [9.0, 8.0, 0.5, 0.4, 1.0]
        market = io.BytesIO()
        with pd.ExcelWriter(market, engine="openpyxl") as writer:
            pd.DataFrame(
                {
                    "\u5de5\u5382": ["BB1"],
                    "\u5206\u7c7b": ["\u817f\u8089"],
                    "\u57fa\u671f\u884c\u60c5\u4ef7": [10.0],
                    "\u5f53\u524d\u884c\u60c5\u4ef7": [12.0],
                }
            ).to_excel(writer, index=False, sheet_name="\u884c\u60c5\u4ef7\u8986\u76d6")
        market.seek(0)

        data = APP.to_excel_bytes(
            leg,
            empty,
            empty,
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame(),
            raw_usage_leg,
            raw_usage_leg.iloc[0:0].copy(),
            raw_usage_leg.iloc[0:0].copy(),
            {raw_code: "\u539f\u6599\u89c4\u683c"},
            {},
            {},
            "\u868c\u57e0\u4e00\u5382\u7cfb\u7edf\u6210\u672c-2604",
            pd.DataFrame(),
            pd.DataFrame(),
            "4\u6708",
            {},
            market,
            quarter_label="Q4",
            market_plant_code="BB1",
        )

        wb = load_workbook(io.BytesIO(data), data_only=False)
        ws = wb["\u817f\u8089\u884c\u60c5-\u8f83\u5b63\u5ea6"]
        headers = [cell.value for cell in ws[2]]
        unit_idx = headers.index(unit_col) + 1
        self.assertEqual(ws.cell(row=5, column=unit_idx).value, 12.0)
        self.assertEqual(ws.cell(row=6, column=unit_idx).value, 10.0)

    def test_market_price_reference_row_stays_zero_when_25_actual_missing(self):
        mat_col = "\u4fee\u884c\u540e\u539f\u6599"
        spec_col = "\u4f7f\u7528\u534a\u6210\u54c1\u89c4\u683c"
        row_col = "\u884c\u7c7b\u578b"
        impact_col = "\u5f71\u54cd\u53e3\u5f84"
        comp_col = "\u7efc\u5408\u5355\u4ef7"
        unit_col = "\u4fee\u5f62\u524d\u539f\u6599\u7efc\u5408\u8017\u7528\u5355\u4ef7"
        util_col = "\u4fee\u5f62\u5229\u7528\u7387"
        loss_col = "\u635f\u8017\u7387"
        raw_cost_col = "\u534a\u6210\u54c1\u539f\u6599\u6210\u672c"
        labor_col = "\u534a\u6210\u54c1\u4fee\u5f62\u4eba\u5de5\u6210\u672c"
        total_col = "\u534a\u6210\u54c1\u603b\u6210\u672c"
        in_qty_col = "\u534a\u6210\u54c1\u5165\u5e93\u91cf"
        cols = [
            mat_col, spec_col, row_col, impact_col, "31000001", comp_col,
            unit_col, util_col, loss_col, raw_cost_col, labor_col, total_col, in_qty_col,
        ]
        header_rows = pd.DataFrame([{c: "" for c in cols} for _ in range(4)])
        data_rows = pd.DataFrame(
            [
                {
                    mat_col: "39000181", spec_col: "\u817f\u6392", row_col: "4\u6708\u5b9e\u9645\u5355\u4ef7", impact_col: "",
                    "31000001": 10.0, comp_col: 10.0, unit_col: 10.0, util_col: 0.6, loss_col: 0.05,
                    raw_cost_col: 12.0, labor_col: 1.0, total_col: 13.0, in_qty_col: 4.0,
                },
                {
                    mat_col: "39000181", spec_col: "\u817f\u6392", row_col: "25\u5e74\u5b9e\u9645\u5355\u4ef7", impact_col: "",
                    "31000001": 0.0, comp_col: 0.0, unit_col: 0.0, util_col: 0.0, loss_col: 0.0,
                    raw_cost_col: 0.0, labor_col: 0.0, total_col: 0.0, in_qty_col: 4.0,
                },
                {mat_col: "39000181", spec_col: "\u817f\u6392", row_col: "4\u6708\u89c4\u683c\u5360\u6bd4", impact_col: "", in_qty_col: 4.0},
                {mat_col: "39000181", spec_col: "\u817f\u6392", row_col: "\u89c4\u683c\u5360\u6bd4", impact_col: "", in_qty_col: 4.0},
                {mat_col: "39000181", spec_col: "\u817f\u6392", row_col: "\u5dee\u5f02", impact_col: "", in_qty_col: 4.0},
                {mat_col: "39000181", spec_col: "\u817f\u6392", row_col: "\u5bf9\u534a\u6210\u54c1\u6210\u672c\u7684\u5f71\u54cd", impact_col: "\u5355\u4f4d\u6210\u672c", in_qty_col: 4.0},
                {mat_col: "39000181", spec_col: "\u817f\u6392", row_col: "\u5bf9\u534a\u6210\u54c1\u6210\u672c\u7684\u5f71\u54cd", impact_col: "\u603b\u6210\u672c", in_qty_col: 4.0},
            ]
        )
        tsc_export = pd.concat([header_rows, data_rows], ignore_index=True).reindex(columns=cols)
        market_map = {
            ("BB2", "\u817f\u8089"): {
                "\u57fa\u671f\u884c\u60c5\u4ef7": 10.0,
                "\u5f53\u524d\u884c\u60c5\u4ef7": 12.0,
            }
        }

        out = APP._build_market_comparison_sheet(tsc_export, "BB2", "\u817f\u8089", "4\u6708", market_map, "Q4")

        self.assertEqual(out.loc[5, row_col], "25\u5e74Q4\u5b9e\u9645\u5355\u4ef7")
        self.assertEqual(out.loc[5, comp_col], 0.0)
        self.assertEqual(out.loc[5, unit_col], 0.0)
        self.assertEqual(out.loc[8, raw_cost_col], 0.0)
        self.assertEqual(out.loc[9, total_col], 0.0)
        self.assertEqual(out.loc[10, total_col], 0.0)

    def test_zero_tsc_reference_metrics_are_treated_as_missing_for_impacts(self):
        zero_metrics = {
            "修形前原料综合耗用单价": 0.0,
            "修形利用率": 0.0,
            "损耗率": 0.0,
            "半成品原料成本": 0.0,
            "半成品修形人工成本": 0.0,
            "半成品总成本": 0.0,
        }
        valid_metrics = dict(zero_metrics)
        valid_metrics["半成品原料成本"] = 12.18

        self.assertFalse(APP._has_tsc_reference_metrics(None))
        self.assertFalse(APP._has_tsc_reference_metrics({}))
        self.assertFalse(APP._has_tsc_reference_metrics(zero_metrics))
        self.assertTrue(APP._has_tsc_reference_metrics(valid_metrics))

    def test_tsc_loss_impact_uses_raw_diff_even_when_display_diff_is_zero(self):
        month_unit = 10.44788802488336
        month_util = 0.6220839813374806
        month_loss = 0.04510108864696732
        q3_loss = 0.0513853289393384
        factor = 0.95
        month_raw_cost = APP._tsc_raw_cost_by(month_unit, month_util, month_loss, factor)

        impact = APP._tsc_loss_impact(
            month_raw_cost,
            month_unit,
            month_util,
            q3_loss,
            factor,
            diff_loss=month_loss - q3_loss,
        )

        self.assertEqual(round(impact, 2), -0.10)

    def test_tsc_price_impact_matches_reference_workbook_price_first_bridge(self):
        cases = [
            {
                "month_unit": 10.20125962732919,
                "reference_unit": 11.472492622704,
                "reference_util": 0.7323095453176754,
                "reference_loss": 0.01565793435712132,
                "reference_raw_cost": 11.91521196178327,
                "expected": -1.3202894166807013,
            },
            {
                "month_unit": 9.750548765533804,
                "reference_unit": 11.23187217221762,
                "reference_util": 0.9067065516412489,
                "reference_loss": 0.04305440559814106,
                "reference_raw_cost": 11.7963276782684,
                "expected": -1.5557670203863339,
            },
        ]

        for case in cases:
            with self.subTest(month_unit=case["month_unit"]):
                impact = APP._tsc_price_impact(
                    case["month_unit"],
                    case["reference_util"],
                    case["reference_loss"],
                    case["reference_raw_cost"],
                    factor=0.95,
                    diff_unit=case["month_unit"] - case["reference_unit"],
                )
                self.assertAlmostEqual(impact, case["expected"], places=12)

    def test_tsc_reference_diff_matches_standard_if_reference_positive_rule(self):
        self.assertEqual(APP._tsc_reference_diff(12.5, 10.0), 2.5)
        self.assertEqual(APP._tsc_reference_diff(12.5, 0.0), 0.0)
        self.assertEqual(APP._tsc_reference_diff(12.5, -0.0), 0.0)
        self.assertEqual(APP._tsc_reference_diff(12.5, -1e-14), 0.0)
        self.assertIsNone(APP._tsc_reference_diff(None, 10.0))
        self.assertIsNone(APP._tsc_reference_diff(12.5, None))

    def test_tsc_reference_costs_are_recalculated_with_current_factor(self):
        metrics = {
            "\u4fee\u5f62\u524d\u539f\u6599\u7efc\u5408\u8017\u7528\u5355\u4ef7": 14.664985,
            "\u4fee\u5f62\u5229\u7528\u7387": 0.463529,
            "\u635f\u8017\u7387": 0.218692,
            "\u534a\u6210\u54c1\u539f\u6599\u6210\u672c": 24.600022,
            "\u534a\u6210\u54c1\u4fee\u5f62\u4eba\u5de5\u6210\u672c": 0.941176,
            "\u534a\u6210\u54c1\u603b\u6210\u672c": 25.541198,
        }

        recalculated = APP._recalculate_tsc_reference_metrics(metrics, factor=0.5)

        self.assertEqual(round(recalculated["\u534a\u6210\u54c1\u539f\u6599\u6210\u672c"], 2), 26.61)
        self.assertEqual(round(recalculated["\u534a\u6210\u54c1\u603b\u6210\u672c"], 2), 27.55)

    def test_tsc_loss_impact_is_zero_when_reference_loss_is_zero(self):
        month_unit = 17.1670167022033
        month_util = 0.635394456289979
        month_loss = 0.000710732054015617
        q3_unit = 18.9348274632189
        q3_loss = -5.55111512312578e-17
        q3_raw_cost = 18.9348274632189
        factor = 1.0
        month_raw_cost = APP._tsc_raw_cost_by(month_unit, month_util, month_loss, factor)
        diff_raw_cost = month_raw_cost - q3_raw_cost
        price_impact = month_raw_cost - APP._tsc_raw_cost_by(q3_unit, month_util, month_loss, factor)
        loss_diff = APP._tsc_loss_diff(month_loss, q3_loss)

        loss_impact = APP._tsc_loss_impact(
            month_raw_cost,
            month_unit,
            month_util,
            q3_loss,
            factor,
            diff_loss=loss_diff,
        )
        util_impact = diff_raw_cost - price_impact - loss_impact

        self.assertEqual(loss_diff, 0.0)
        self.assertEqual(round(loss_impact, 2), 0.00)
        self.assertEqual(round(util_impact, 2), 0.02)

    def test_tsc_loss_impact_uses_one_when_material_factor_is_zero(self):
        month_unit = 37.745016
        month_util = 0.93949
        month_loss = 0.06051
        q3_unit = 34.862397
        q3_util = 0.924793
        q3_loss = 0.075207
        q3_raw_cost = 37.697497
        month_raw_cost = APP._tsc_raw_cost_by(month_unit, month_util, month_loss, 0.0)
        diff_raw_cost = month_raw_cost - q3_raw_cost
        price_impact = month_raw_cost - APP._tsc_raw_cost_by(q3_unit, month_util, month_loss, 0.0)
        loss_impact = APP._tsc_loss_impact(
            month_raw_cost,
            month_unit,
            month_util,
            q3_loss,
            APP._tsc_loss_impact_factor(0.0),
            diff_loss=month_loss - q3_loss,
        )
        util_impact = diff_raw_cost - price_impact - loss_impact

        self.assertEqual(round(price_impact, 2), 3.07)
        self.assertEqual(round(util_impact, 2), 0.00)
        self.assertEqual(round(loss_impact, 2), -0.59)

    def test_tsc_preview_formats_ratio_columns_by_row_type(self):
        row_type_col = "\u884c\u7c7b\u578b"
        impact_col = "\u5f71\u54cd\u53e3\u5f84"
        util_col = "\u4fee\u5f62\u5229\u7528\u7387"
        loss_col = "\u635f\u8017\u7387"
        df = pd.DataFrame(
            [
                {row_type_col: "4\u6708\u5b9e\u9645\u5355\u4ef7", impact_col: "", util_col: 0.64, loss_col: 0.0},
                {row_type_col: "25Q4\u5b9e\u9645\u5355\u4ef7", impact_col: "", util_col: 0.64, loss_col: -0.0},
                {row_type_col: "10\u6708\u89c4\u683c\u5360\u6bd4", impact_col: "", util_col: None, loss_col: None},
                {row_type_col: "\u5dee\u5f02", impact_col: "", util_col: -0.0, loss_col: 0.0},
                {row_type_col: "\u5bf9\u534a\u6210\u54c1\u6210\u672c\u7684\u5f71\u54cd", impact_col: "\u5355\u4f4d\u6210\u672c", util_col: 0.02, loss_col: 0.0},
                {row_type_col: "\u5bf9\u534a\u6210\u54c1\u6210\u672c\u7684\u5f71\u54cd", impact_col: "\u603b\u6210\u672c", util_col: 0.08, loss_col: 0.0},
            ]
        )

        display = APP._format_tsc_display(df)

        self.assertEqual(display.loc[0, util_col], "64%")
        self.assertEqual(display.loc[1, loss_col], "0%")
        self.assertEqual(display.loc[2, util_col], "")
        self.assertEqual(display.loc[3, util_col], "0%")
        self.assertEqual(display.loc[4, util_col], "0.02")
        self.assertEqual(display.loc[4, loss_col], "0.00")
        self.assertEqual(display.loc[5, util_col], "0.08")

    def test_tsc_zero_loss_export_formats_show_zero_not_dash(self):
        tsc_rows = pd.DataFrame(
            [
                {
                    "物料号": "39000340",
                    "物料描述": "猪4号肉片/6-15g/自修形",
                    "行类型": "4月实际单价",
                    "影响口径": "",
                    "分类": "锅包肉",
                    "修形前原料综合耗用单价": 16.13,
                    "修形利用率": 0.70,
                    "损耗率": 0.0,
                    "半成品原料成本": 16.13,
                    "半成品修形人工成本": 0.83,
                    "半成品总成本": 16.96,
                    "半成品入库量": 1.84,
                    "BOM": "",
                    "BOM占比": None,
                },
                {
                    "物料号": "39000340",
                    "物料描述": "猪4号肉片/6-15g/自修形",
                    "行类型": "差异",
                    "影响口径": "",
                    "分类": "锅包肉",
                    "修形前原料综合耗用单价": -1.02,
                    "修形利用率": 0.04,
                    "损耗率": 0.0,
                    "半成品原料成本": -0.98,
                    "半成品修形人工成本": -0.79,
                    "半成品总成本": -1.77,
                    "半成品入库量": 1.84,
                    "BOM": "",
                    "BOM占比": None,
                },
                {
                    "物料号": "39000340",
                    "物料描述": "猪4号肉片/6-15g/自修形",
                    "行类型": "对半成品成本的影响",
                    "影响口径": "单位成本",
                    "分类": "锅包肉",
                    "修形前原料综合耗用单价": -1.02,
                    "修形利用率": 0.04,
                    "损耗率": 0.0,
                    "半成品原料成本": -0.98,
                    "半成品修形人工成本": -0.79,
                    "半成品总成本": -1.77,
                    "半成品入库量": 1.84,
                    "BOM": "",
                    "BOM占比": None,
                },
            ]
        )
        raw_usage = tsc_rows[["物料号", "物料描述", "行类型", "分类"]].copy()
        empty_tsc = tsc_rows.iloc[0:0].copy()
        empty_raw_usage = raw_usage.iloc[0:0].copy()

        data = APP.to_excel_bytes(
            tsc_rows,
            empty_tsc,
            empty_tsc,
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame(),
            raw_usage,
            empty_raw_usage,
            empty_raw_usage,
            {},
            {},
            {},
            "4月",
            pd.DataFrame(),
            pd.DataFrame(),
            "4月",
            {},
        )

        wb = load_workbook(io.BytesIO(data))
        ws = wb["腿肉TSC"]
        headers = [cell.value for cell in ws[2]]
        util_col = headers.index("修形利用率") + 1
        loss_col = headers.index("损耗率") + 1

        self.assertEqual(ws.cell(row=5, column=loss_col).number_format, "0%;[Red](0%);0%")
        self.assertEqual(ws.cell(row=6, column=loss_col).number_format, "0.0%;[Red](0.0%);0.0%")
        self.assertEqual(ws.cell(row=7, column=util_col).number_format, "0.00;[Red](0.00);0.00")
        self.assertEqual(ws.cell(row=7, column=loss_col).number_format, "0.00;[Red](0.00);0.00")

    def test_ensure_and_map_columns_uses_aliases_and_clean_names(self):
        src = pd.DataFrame(
            columns=[
                " 物料编号 ",
                "品名",
                "原料编码",
                "原料名称",
                "入库数量（kg）",
                "入库金额含税",
                "实际量",
                "实际金额（元）",
                "配方用量",
            ]
        )

        out = APP._ensure_and_map_columns(
            src,
            ["物料号", "物料描述", "原料号", "原料描述", "入库数量", "入库金额", "实际数量", "实际金额", "配方数量"],
        )

        self.assertEqual(
            list(out.columns),
            ["物料号", "物料描述", "原料号", "原料描述", "入库数量", "入库金额", "实际数量", "实际金额", "配方数量"],
        )

    def test_ensure_and_map_columns_falls_back_to_material_keyword(self):
        src = pd.DataFrame(
            columns=["物料代码", "物料名称", "原料", "原料品名", "入库量", "入库金额(元)", "实际数量(kg)", "实际金额含税", "配方数量（kg）"]
        )

        out = APP._ensure_and_map_columns(
            src,
            ["物料号", "物料描述", "原料号", "原料描述", "入库数量", "入库金额", "实际数量", "实际金额", "配方数量"],
        )

        self.assertIn("物料号", out.columns)
        self.assertEqual(out.columns[0], "物料号")

    def test_month_and_quarter_labels_from_filename(self):
        self.assertEqual(APP._month_label_from_filename("BB2_2601.xlsx"), "1月")
        self.assertEqual(APP._month_label_from_filename("BB2_11.xlsx"), "11月")
        self.assertEqual(APP._month_label_from_filename("plain-name.xlsx"), "11月")
        self.assertEqual(APP._quarter_label_from_filename("成本25Q2-Q4.xlsx"), "Q2")
        self.assertEqual(APP._quarter_label_from_filename("no-quarter.xlsx"), "Q3")

    def test_month_code_and_output_filename(self):
        self.assertEqual(APP._month_code_from_filename("BB2_2601.xlsx"), "2601")
        self.assertEqual(APP._month_code_from_filename("report_01.xlsx"), "1")
        self.assertEqual(
            APP._build_output_filename("BB2_2601.xlsx", "原料清单_蚌埠二厂.xlsx"),
            "蚌埠二厂系统成本-2601.xlsx",
        )
        self.assertEqual(
            APP._build_output_filename("compare.xlsx", "原料清单_天津.xlsx"),
            "天津11月系统成本.xlsx",
        )

    def test_resolve_tsc_quarter_label_prefers_requested_quarter_when_present(self):
        leg = pd.DataFrame([[None, None, None, None, "Q2实际单价"], [None, None, None, None, "Q2规格占比"]])
        bre = pd.DataFrame([[None, None, None, None, "Q1实际单价"], [None, None, None, None, "Q2实际单价"]])
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            leg.to_excel(writer, sheet_name="腿肉TSC", header=False, index=False)
            bre.to_excel(writer, sheet_name="胸肉TSC", header=False, index=False)
        buf.seek(0)

        self.assertEqual(APP._resolve_tsc_quarter_label(buf, preferred="Q2"), "Q2")

    def test_resolve_tsc_quarter_label_falls_back_to_most_frequent_quarter(self):
        leg = pd.DataFrame(
            [
                [None, None, None, None, "Q1实际单价"],
                [None, None, None, None, "Q1规格占比"],
                [None, None, None, None, "Q1实际单价"],
            ]
        )
        bre = pd.DataFrame([[None, None, None, None, "Q3实际单价"]])
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            leg.to_excel(writer, sheet_name="腿肉TSC", header=False, index=False)
            bre.to_excel(writer, sheet_name="胸肉TSC", header=False, index=False)
        buf.seek(0)

        self.assertEqual(APP._resolve_tsc_quarter_label(buf, preferred="Q4"), "Q1")

    def test_tsc_reference_lookup_uses_header_names_not_fixed_positions(self):
        sheet_name = "\u817f\u8089TSC"
        mat_col = "\u4fee\u884c\u540e\u539f\u6599"
        label_col = "\u884c\u7c7b\u578b"
        impact_col = "\u5f71\u54cd\u53e3\u5f84"
        unit_col = "\u4fee\u5f62\u524d\u539f\u6599\u7efc\u5408\u8017\u7528\u5355\u4ef7"
        util_col = "\u4fee\u5f62\u5229\u7528\u7387"
        loss_col = "\u635f\u8017\u7387"
        raw_cost_col = "\u534a\u6210\u54c1\u539f\u6599\u6210\u672c"
        labor_col = "\u534a\u6210\u54c1\u4fee\u5f62\u4eba\u5de5\u6210\u672c"
        total_col = "\u534a\u6210\u54c1\u603b\u6210\u672c"
        comp_col = "\u7efc\u5408\u5355\u4ef7"
        df = pd.DataFrame(
            [
                [
                    "\u4e0d\u76f8\u5173", loss_col, label_col, total_col, util_col,
                    mat_col, raw_cost_col, unit_col, labor_col, impact_col, comp_col,
                ],
                [
                    "", 0.22, "25\u5e74\u5b9e\u9645\u5355\u4ef7", 27.55, 0.46,
                    "39001158", 26.61, 14.66, 0.94, "", 14.66,
                ],
                [
                    "", None, "Q4\u89c4\u683c\u5360\u6bd4", None, None,
                    "39001158", None, None, None, "", 0.85,
                ],
            ]
        )
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
        buf.seek(0)

        metrics = APP._find_tsc_metrics(buf, sheet_name, "39001158", "25\u5e74\u5b9e\u9645\u5355\u4ef7")
        buf.seek(0)
        ratio = APP._find_tsc_value(buf, sheet_name, "39001158", "Q4\u89c4\u683c\u5360\u6bd4")

        self.assertEqual(metrics[unit_col], 14.66)
        self.assertEqual(metrics[util_col], 0.46)
        self.assertEqual(metrics[loss_col], 0.22)
        self.assertEqual(metrics[raw_cost_col], 26.61)
        self.assertEqual(metrics[labor_col], 0.94)
        self.assertEqual(metrics[total_col], 27.55)
        self.assertEqual(ratio, 0.85)

    def test_quarter_reference_metrics_match_25q_label(self):
        sheet_name = "\u80f8\u8089TSC"
        mat_col = "\u4fee\u884c\u540e\u539f\u6599"
        label_col = "\u884c\u7c7b\u578b"
        unit_col = "\u4fee\u5f62\u524d\u539f\u6599\u7efc\u5408\u8017\u7528\u5355\u4ef7"
        util_col = "\u4fee\u5f62\u5229\u7528\u7387"
        loss_col = "\u635f\u8017\u7387"
        raw_cost_col = "\u534a\u6210\u54c1\u539f\u6599\u6210\u672c"
        labor_col = "\u534a\u6210\u54c1\u4fee\u5f62\u4eba\u5de5\u6210\u672c"
        total_col = "\u534a\u6210\u54c1\u603b\u6210\u672c"
        df = pd.DataFrame(
            [
                [mat_col, label_col, unit_col, util_col, loss_col, raw_cost_col, labor_col, total_col],
                ["39000392", "25\u5e74\u5b9e\u9645\u5355\u4ef7", 8.51, 0.34, 0.06, 14.33, 1.51, 15.84],
                ["39000392", "25Q4\u5b9e\u9645\u5355\u4ef7", 7.59, 0.33, 0.06, 13.30, 1.31, 14.61],
            ]
        )
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
        buf.seek(0)

        metrics = APP._find_tsc_reference_metrics(buf, sheet_name, "39000392", "25\u5e74\u5b9e\u9645\u5355\u4ef7", "Q4")

        self.assertEqual(metrics[unit_col], 7.59)
        self.assertTrue(APP._has_tsc_reference_metrics(metrics))

    def test_tsc_main_reference_metrics_can_keep_25_year_label(self):
        sheet_name = "\u80f8\u8089TSC"
        mat_col = "\u4fee\u884c\u540e\u539f\u6599"
        label_col = "\u884c\u7c7b\u578b"
        unit_col = "\u4fee\u5f62\u524d\u539f\u6599\u7efc\u5408\u8017\u7528\u5355\u4ef7"
        util_col = "\u4fee\u5f62\u5229\u7528\u7387"
        loss_col = "\u635f\u8017\u7387"
        raw_cost_col = "\u534a\u6210\u54c1\u539f\u6599\u6210\u672c"
        labor_col = "\u534a\u6210\u54c1\u4fee\u5f62\u4eba\u5de5\u6210\u672c"
        total_col = "\u534a\u6210\u54c1\u603b\u6210\u672c"
        df = pd.DataFrame(
            [
                [mat_col, label_col, unit_col, util_col, loss_col, raw_cost_col, labor_col, total_col],
                ["39000392", "25\u5e74\u5b9e\u9645\u5355\u4ef7", 8.51, 0.34, 0.06, 14.33, 1.51, 15.84],
                ["39000392", "25Q4\u5b9e\u9645\u5355\u4ef7", 7.59, 0.33, 0.06, 13.30, 1.31, 14.61],
            ]
        )
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
        buf.seek(0)

        metrics = APP._find_tsc_reference_metrics(
            buf,
            sheet_name,
            "39000392",
            "25\u5e74\u5b9e\u9645\u5355\u4ef7",
            "Q4",
            prefer_quarter=False,
        )

        self.assertEqual(metrics[unit_col], 8.51)
        self.assertEqual(metrics[raw_cost_col], 14.33)

    def test_quarter_reference_raw_values_match_25q_label(self):
        df = pd.DataFrame(
            [
                ["", "", "31000001", "\u7efc\u5408\u5355\u4ef7"],
                [
                    "\u884c\u7c7b\u578b",
                    "\u4fee\u884c\u540e\u539f\u6599",
                    "",
                    "",
                ],
                ["25\u5e74\u5b9e\u9645\u5355\u4ef7", "39000392", 15.35, 15.35],
                ["25Q4\u5b9e\u9645\u5355\u4ef7", "39000392", 14.61, 14.61],
            ]
        )

        values = APP._find_tsc_reference_row_values(
            df,
            "39000392",
            "25\u5e74\u5b9e\u9645\u5355\u4ef7",
            "Q4",
            [(2, "31000001")],
            3,
        )

        self.assertEqual(values["31000001"], 14.61)
        self.assertEqual(values["\u7efc\u5408\u5355\u4ef7"], 14.61)

    def test_tsc_main_reference_raw_values_can_keep_25_year_label(self):
        df = pd.DataFrame(
            [
                ["", "", "31000001", "\u7efc\u5408\u5355\u4ef7"],
                [
                    "\u884c\u7c7b\u578b",
                    "\u4fee\u884c\u540e\u539f\u6599",
                    "",
                    "",
                ],
                ["25\u5e74\u5b9e\u9645\u5355\u4ef7", "39000392", 15.35, 15.35],
                ["25Q4\u5b9e\u9645\u5355\u4ef7", "39000392", 14.61, 14.61],
            ]
        )

        values = APP._find_tsc_reference_row_values(
            df,
            "39000392",
            "25\u5e74\u5b9e\u9645\u5355\u4ef7",
            "Q4",
            [(2, "31000001")],
            3,
            prefer_quarter=False,
        )

        self.assertEqual(values["31000001"], 15.35)
        self.assertEqual(values["\u7efc\u5408\u5355\u4ef7"], 15.35)

    def test_tsc_raw_columns_use_header_fields_when_columns_are_reordered(self):
        sheet_name = "\u817f\u8089TSC"
        df = pd.DataFrame(
            [
                [
                    "\u884c\u7c7b\u578b",
                    "\u4fee\u884c\u540e\u539f\u6599",
                    "31000001",
                    "\u7efc\u5408\u5355\u4ef7",
                    "31000002",
                    "\u4fee\u5f62\u524d\u539f\u6599\u7efc\u5408\u8017\u7528\u5355\u4ef7",
                ],
                ["25\u5e74\u5b9e\u9645\u5355\u4ef7", "39001158", 1.2, 14.66, 2.3, 14.66],
            ]
        )
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
        buf.seek(0)

        raw_cols, comp_col, _, _ = APP._get_tsc_raw_columns(buf, sheet_name)

        self.assertEqual(raw_cols, [(2, "31000001"), (4, "31000002")])
        self.assertEqual(comp_col, 3)

    def test_filter_parts_by_material_spec_keeps_semifinished_when_any_raw_hits_whitelist(self):
        parts = [
            {"修行后原料": "39000134", "原料号": "31001602", "数量": 10},
            {"修行后原料": "39000134", "原料号": "31002579", "数量": 5},
            {"修行后原料": "39000134", "原料号": "31003191", "数量": 3},
            {"修行后原料": "39000223", "原料号": "31009999", "数量": 8},
        ]
        profile = {
            "lookup": {
                "胸肉": {
                    "31001602": "无规格",
                    "31003191": "220-260g",
                }
            }
        }

        filtered, allowed_mats = FRESH_APP._filter_parts_by_material_spec("胸肉", parts, profile)

        self.assertEqual(
            [part["原料号"] for part in filtered],
            ["31001602", "31003191"],
        )
        self.assertEqual(allowed_mats, {"39000134"})

    def test_filter_parts_by_material_spec_without_profile_keeps_all_parts(self):
        parts = [
            {"修行后原料": "39000134", "原料号": "31001602", "数量": 10},
            {"修行后原料": "39000223", "原料号": "31009999", "数量": 8},
        ]

        filtered, allowed_mats = FRESH_APP._filter_parts_by_material_spec("胸肉", parts, None)

        self.assertEqual(filtered, parts)
        self.assertEqual(allowed_mats, {"39000134", "39000223"})


if __name__ == "__main__":
    unittest.main()
